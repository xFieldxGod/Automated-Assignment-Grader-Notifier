"""
bot1_data_processor.py
------------------------------------------
Bot 1 - Data Processor
หน้าที่: อ่านข้อมูลนักศึกษา + คะแนน แล้วคืนเป็น list[dict] ให้บอทอื่นใช้ต่อ

รองรับ 2 แหล่ง:
  1) Excel (.xlsx)                    - ใช้ openpyxl
  2) Google Sheets (Form Responses)   - ใช้ gspread

รูปแบบข้อมูลที่คืน (ต่อ 1 นักศึกษา):
{
    "student_id": "6400001",
    "name": "สมชาย ใจดี",
    "email": "somchai@example.com",
    "scores": {"การบ้าน 1": 9, "การบ้าน 2": 8, ...},
    "total": 85,
    "full_total": 100,
    "percent": 85.0,
    "grade": "A",
}
"""
from __future__ import annotations

# เพิ่มโฟลเดอร์แม่เข้า sys.path ให้ `import config` ทำงานได้ทั้งตอนรันผ่าน main.py
# และตอนรันไฟล์นี้เดี่ยว ๆ (py bots/bot1_data_processor.py)
import sys
from pathlib import Path
_PARENT = Path(__file__).resolve().parent.parent
if str(_PARENT) not in sys.path:
    sys.path.insert(0, str(_PARENT))

import logging
from typing import Iterator

import config

log = logging.getLogger("bot1")


# ---------------------------------------------------------
# ฟังก์ชันคำนวณ (ใช้ได้กับทั้งสองแหล่งข้อมูล)
# ---------------------------------------------------------
def _calc_grade(percent: float) -> str:
    """แปลงเปอร์เซ็นต์เป็นเกรดตามตารางใน config.GRADE_TABLE"""
    for threshold, grade in config.GRADE_TABLE:
        if percent >= threshold:
            return grade
    return "F"


def _build_record(raw: dict) -> dict:
    """รวมข้อมูลดิบ (dict จาก 1 แถว) ให้อยู่ในรูปแบบมาตรฐาน"""
    full_total = sum(config.ASSIGNMENTS.values())

    scores: dict[str, float] = {}
    total = 0.0
    for col, full in config.ASSIGNMENTS.items():
        value = raw.get(col, 0) or 0
        try:
            score = float(value)
        except (TypeError, ValueError):
            log.warning("  แปลงคะแนน %r ของคอลัมน์ %r ไม่ได้ ใช้ 0 แทน", value, col)
            score = 0.0
        # ป้องกันใส่เกินคะแนนเต็ม
        score = max(0.0, min(score, full))
        scores[col] = score
        total += score

    percent = round((total / full_total) * 100, 2) if full_total else 0.0

    return {
        "student_id": str(raw.get(config.COL_STUDENT_ID, "")).strip(),
        "name": str(raw.get(config.COL_NAME, "")).strip(),
        "email": str(raw.get(config.COL_EMAIL, "")).strip(),
        "scores": scores,
        "total": round(total, 2),
        "full_total": full_total,
        "percent": percent,
        "grade": _calc_grade(percent),
    }


# ---------------------------------------------------------
# อ่านจาก Excel
# ---------------------------------------------------------
def _rows_from_excel(path) -> Iterator[dict]:
    from openpyxl import load_workbook

    log.info("อ่านข้อมูลจากไฟล์ Excel: %s", path)
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    for row in rows:
        if row is None or all(v is None for v in row):
            continue  # ข้ามบรรทัดว่าง
        yield dict(zip(headers, row))


# ---------------------------------------------------------
# อ่านจาก Google Sheets
# ---------------------------------------------------------
def _rows_from_gsheet() -> Iterator[dict]:
    import gspread
    from google.oauth2.service_account import Credentials

    log.info("อ่านข้อมูลจาก Google Sheet: %s", config.GSHEET_SPREADSHEET_ID)
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    creds = Credentials.from_service_account_file(
        config.GSHEET_CREDENTIALS, scopes=scopes
    )
    client = gspread.authorize(creds)
    sh = client.open_by_key(config.GSHEET_SPREADSHEET_ID)
    ws = sh.worksheet(config.GSHEET_WORKSHEET_NAME)

    # get_all_records() คืน list[dict] แถวละ dict โดยใช้ header แถวแรกเป็น key
    for record in ws.get_all_records():
        yield record


# ---------------------------------------------------------
# จุดเข้าใช้งานหลัก
# ---------------------------------------------------------
def load_students() -> list[dict]:
    """คืน list[dict] ของนักศึกษา พร้อมคะแนนที่คำนวณแล้ว"""
    if config.SOURCE == "excel":
        rows = _rows_from_excel(config.EXCEL_FILE)
    elif config.SOURCE == "gsheet":
        rows = _rows_from_gsheet()
    else:
        raise ValueError(f"ไม่รู้จัก SOURCE={config.SOURCE!r} (ต้องเป็น 'excel' หรือ 'gsheet')")

    students: list[dict] = []
    for raw in rows:
        rec = _build_record(raw)
        # ข้ามแถวที่ไม่มีรหัสหรือไม่มีอีเมล
        if not rec["student_id"] or not rec["email"]:
            log.warning("ข้ามแถวที่ข้อมูลไม่ครบ: %r", raw)
            continue
        students.append(rec)

    log.info("โหลดข้อมูลนักศึกษาได้ %d คน", len(students))
    return students


if __name__ == "__main__":
    # ทดสอบ Bot 1 เดี่ยว ๆ: python bot1_data_processor.py
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(name)s] %(message)s")
    data = load_students()
    for s in data:
        print(f"{s['student_id']}  {s['name']:<25}  "
              f"{s['total']}/{s['full_total']}  ({s['percent']}%)  เกรด {s['grade']}")
